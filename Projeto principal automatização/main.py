from openpyxl import load_workbook
import time
import schedule

def auto_update(planilha_original):
    arquivo = load_workbook(planilha_original)
    aba_utilizada = arquivo.active
    Cell = int(input("Qual linha desejada: "))
    Cell_column = int(input("Qual a coluna desejada: "))
    valor_Cell = aba_utilizada.cell(row=Cell, column=Cell_column).value
    valor_Cell = aba_utilizada.cell(row=Cell, column=Cell_column).value = valor_Cell + valor_Cell*2/100
    arquivo.save("update.xlsx")

auto_update("C:\Desktop\Projeto principal automatização\pLA_aUTO.xlsx")