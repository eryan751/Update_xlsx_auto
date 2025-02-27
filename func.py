from openpyxl import load_workbook
import pandas as pd
import os
import smtplib
from email.message import EmailMessage
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart

def update_sheet(planilha_original):
    arquivo = load_workbook(planilha_original)
    aba_utilizada = arquivo.active

    while True:
        confirmação = str(input("Quer alterar alguma celula? (S/N) ")).upper()
        if confirmação == "S":
            while True:  # Loop interno para garantir entradas válidas
                try:
                    linha = int(input("Qual a linha desejada: "))
                    coluna = int(input("Qual a coluna desejada: "))
                    valor_aumento = float(input("Em quantos % deseja aumentar? "))
                    break  # Sai do loop interno se as entradas forem válidas
                except ValueError:
                    print("Entrada inválida. Certifique-se de inserir números inteiros para linha e coluna e um número para a porcentagem.")

            try:
                valor_Cell = aba_utilizada.cell(row=linha, column=coluna).value
                if isinstance(valor_Cell, (int, float)):  # Verifica se o valor da célula é numérico
                    novo_valor = valor_Cell * (1 + valor_aumento / 100)  # Calcula o novo valor com base na porcentagem
                    aba_utilizada.cell(row=linha, column=coluna).value = novo_valor
                    print("Celula atualizada!")
                else:
                    print("A célula selecionada não contém um valor numérico.")
            except IndexError:
                print("Linha ou coluna fora do intervalo da planilha.")

            alteração = str(input("Deseja alterar outra célula? (S/N) ")).upper()
            if alteração != "S":
                break  # Sai do loop principal se o usuário não quiser alterar mais células
        elif confirmação == "N":
            print("Ok, email em breve será ENVIADO...")
            break  # Sai do loop principal se o usuário não quiser fazer alterações
        else:
            print("Opção inválida. Digite 'S' para sim ou 'N' para não.")

    arquivo.save("dados_finais.xlsx")


#Adicionar tratamento de erros
def process_send():
    # Lendo e atualizando dados
    dados = pd.read_excel("Reservar 2.xlsx", header=2)
    dados['Valor Total'] = dados['Preço Unitário'] * 0.1 * dados['Qtd_vendida']
    dados_utilizados = ["Produtos", "Qtd_vendida", "Valor Total"]
    df = dados[dados_utilizados]

    # Adicionando a linha de totais
    totais = pd.DataFrame({'Produtos': ['Totais'], 
                           'Qtd_vendida': [df['Qtd_vendida'].sum()], 
                           'Valor Total': [df['Valor Total'].sum()]})
    df = pd.concat([df, totais], ignore_index=True)
    # Removendo a linha 44 (se existir)
    if 44 in df.index:
        df.drop(44, inplace=True)

    # Removendo valores ausentes (se necessário)
    df.dropna(inplace=True)
    print("Alteraçoes concluidas")

    # Salvando a nova versão
    df.to_excel('dados_finais.xlsx', index=False)


    #Definindo o email remetente
    EMAIL_ADDRESS = 'emanueler8010@gmail.com'
    #Senha gerada para este app em especifico
    EMAIL_PASSWORD = 'mjmxfiyjszawxdrr'

    #Assunto e email destinatario
    msg = MIMEMultipart()
    msg['Subject'] = "Fatutamento do Mês"
    msg['From'] = 'emanueler8010@gmail.com'
    msg['To'] = 'emadevautonext@gmail.com'
    msg.attach(EmailMessage("Faturamento deste Mês"))

    #Escolhendo e lendo Anexo que será enviao
    caminho_arquivo = "dados_finais.xlsx"
    with open(caminho_arquivo, 'rb') as arquivo:
        parte = MIMEBase('application', 'octet-stream')
        parte.set_payload(arquivo.read())

    encoders.encode_base64(parte)
    parte.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(caminho_arquivo)}"')
    msg.attach(parte)

    #Enviando
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(EMAIL_ADDRESS,EMAIL_PASSWORD)
        smtp.send_message(msg)
    
    print("Email enviado")
