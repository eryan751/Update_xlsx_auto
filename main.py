from openpyxl import load_workbook
import time
import schedule

def auto_update(planilha_original, row, column):
    arquivo = load_workbook(planilha_original)
    aba_utilizada = arquivo.active
    valor_Cell = aba_utilizada.cell(row=row, column=column).value
    novo_valor = valor_Cell * 1.02
    aba_utilizada.cell(row=row, column=column).value = novo_valor
    arquivo.save("update.xlsx")
            
linha = int(input("Qual linha desejada: "))
coluna = int(input("Qual a coluna desejada: "))

#Agendando a função
schedule.every().day.at("14:02:30").do(auto_update, planilha_original="update.xlsx", row=linha, column=coluna)

while True:
    schedule.run_pending()
